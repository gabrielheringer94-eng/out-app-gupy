"""
Preprocessor para o piloto Fabbrica.
Lê os arquivos da pasta SEX 10, aplica forward-fill + filtro de subtotal,
filtra ao allowlist de BDRs e gera um JSON flat embutível no HTML.

bdr_col e ffill_cols são resolvidos dinamicamente pelos NOMES de coluna,
porque exports SF têm layouts variáveis (col 0 às vezes é vazio).
"""
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
import csv as csv_mod

DATA_DIR = Path("/Users/gabrielferreiraheringer/Desktop/SEX 10")
OUT_PATH = Path(__file__).parent / "_piloto_data.json"

BDR_ALLOWLIST = [
    {"match": "ariane",         "display": "Ariane Mesquita",   "grupo": "Controle"},
    {"match": "juliana sant",   "display": "Juliana Sant'Anna", "grupo": "Controle"},
    {"match": "beatriz souza",  "display": "Beatriz Souza",     "grupo": "Controle"},
    {"match": "pedro",          "display": "Pedro",             "grupo": "Teste"},
    {"match": "jeanne",         "display": "Jeanne",            "grupo": "Teste"},
    {"match": "victoria",       "display": "Victoria",          "grupo": "Teste"},
    {"match": "mydoha",         "display": "Mydoha Seydo",      "grupo": "Teste"},
    {"match": "giovana",        "display": "Giovana",           "grupo": "Teste"},
    {"match": "adhara",         "display": "Adhara Gonçalves",  "grupo": "Teste"},
]

def classify(raw_name):
    if not raw_name:
        return None
    s = str(raw_name).strip().lower()
    if not s or s in ("subtotal", "total geral"):
        return None
    for e in BDR_ALLOWLIST:
        if e["match"] in s:
            return e
    return None

def to_iso_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return f"{y:04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})", s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None

def find_header_idx(grid, tokens):
    for i in range(min(len(grid), 30)):
        row = grid[i] or []
        text = " | ".join(str(c) if c is not None else "" for c in row)
        if all(tok in text for tok in tokens):
            return i
    return -1

def normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"\s+[↑↓]$", "", str(h)).strip()

def find_col(headers, regex):
    for i, h in enumerate(headers):
        if re.search(regex, h or "", re.IGNORECASE):
            return i
    return -1

def find_cols_all(headers, regexes):
    return [find_col(headers, r) for r in regexes]

def read_xlsx_grid(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))
    return grid

def is_subtotal_row(row, ff_cols):
    for c in ff_cols:
        if 0 <= c < len(row):
            v = row[c]
            if v is None:
                continue
            s = str(v).strip().lower()
            if s in ("subtotal", "total geral"):
                return True
    return False

def apply_ffill(row, ff, ff_cols):
    new_row = list(row)
    max_col = max(ff_cols) if ff_cols else -1
    if max_col >= len(new_row):
        new_row.extend([None] * (max_col + 1 - len(new_row)))
    for k, col in enumerate(ff_cols):
        v = new_row[col]
        if v is not None and str(v).strip() != "":
            ff[col] = str(v).strip()
            for kk in range(k + 1, len(ff_cols)):
                ff[ff_cols[kk]] = ""
    for col in ff_cols:
        v = new_row[col]
        if v is None or str(v).strip() == "":
            new_row[col] = ff.get(col, "")
    return new_row

def parse_xlsx(path, schema, debug=False):
    grid = read_xlsx_grid(path)
    hdr_idx = find_header_idx(grid, schema["header_tokens"])
    if hdr_idx < 0:
        raise RuntimeError(f"Header não encontrado em {path.name}")
    headers = [normalize_header(h) for h in grid[hdr_idx]]
    cols = schema["resolve_cols"](headers)

    bdr_col = find_col(headers, schema["bdr_col_re"])
    ffill_cols = [c for c in find_cols_all(headers, schema["ffill_res"]) if c >= 0]

    if bdr_col < 0:
        raise RuntimeError(f"Coluna BDR não encontrada em {path.name} · pattern={schema['bdr_col_re']}")

    if debug:
        print(f"  hdr_idx={hdr_idx} bdr_col={bdr_col} ffill={ffill_cols} cols={cols}")

    ff = {c: "" for c in ffill_cols}
    records = []
    counts = {"subtotal": 0, "no_bdr": 0, "extract_none": 0, "kept": 0}
    for i in range(hdr_idx + 1, len(grid)):
        row = grid[i] or []
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        filled = apply_ffill(row, ff, ffill_cols)
        if is_subtotal_row(filled, ffill_cols):
            counts["subtotal"] += 1; continue
        bdr_raw = filled[bdr_col] if bdr_col < len(filled) else None
        cls = classify(bdr_raw)
        if not cls:
            counts["no_bdr"] += 1; continue
        rec = schema["extract"](filled, cols, cls)
        if rec is None:
            counts["extract_none"] += 1; continue
        counts["kept"] += 1
        records.append(rec)
    if debug:
        print(f"  counts={counts}")
    return records

def get_str(row, idx):
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()

def get_iso(row, idx):
    if idx < 0 or idx >= len(row):
        return None
    return to_iso_date(row[idx])

# ============================================================
# SCHEMAS
# ============================================================
SCHEMAS = {}

SCHEMAS["atividadesConcluidas"] = {
    "filename_glob": "Atividades concluídas*.xlsx",
    "header_tokens": ["Subtipo de Tarefa", "Criado por: Nome completo"],
    "bdr_col_re": r"Criado por: Nome completo",
    "ffill_res": [r"Subtipo de Tarefa", r"^Assunto$", r"Criado por: Nome completo"],
    "resolve_cols": lambda h: {
        "data": find_col(h, r"Data e hora de conclusão"),
        "conta": find_col(h, r"Conta: Nome da conta"),
        "tipo": find_col(h, r"^Tipo$"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "d": get_iso(row, c["data"]),
            "c": get_str(row, c["conta"]) or None,
            "t": get_str(row, c["tipo"]) or None,
        }
        if get_str(row, c["data"]) and get_str(row, c["tipo"])
        else None
    ),
}

SCHEMAS["ligacoesConectadas"] = {
    "filename_glob": "Ligações Conectadas*.xlsx",  # without [Pre-Vendas] prefix
    "header_tokens": ["Responsável: Nome completo", "Status da Conexão"],
    "bdr_col_re": r"Responsável: Nome completo",
    "ffill_res": [r"Responsável: Nome completo"],
    "resolve_cols": lambda h: {
        "conta": find_col(h, r"Conta: Nome da conta"),
        "data": find_col(h, r"Data de criação"),
        "status_conexao": find_col(h, r"Status da Conexão"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "d": get_iso(row, c["data"]),
            "c": get_str(row, c["conta"]) or None,
        }
        if get_str(row, c["data"]) and get_str(row, c["conta"])
        else None
    ),
}

SCHEMAS["leadsNovo"] = {
    "filename_glob": "Leads por Etapa - Novo Lead*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Etapa do Lead"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe", r"ID da conta"],
    "resolve_cols": lambda h: {
        "negocio": find_col(h, r"Nome do negócio"),
        "etapa_lead": find_col(h, r"Etapa do Lead"),
        "etapa_neg": find_col(h, r"Etapa do negócio"),
        "id_conta": find_col(h, r"ID da conta"),
        "data": find_col(h, r"Data de entrada"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "n": get_str(row, c["negocio"]),
            "el": get_str(row, c["etapa_lead"]),
            "en": get_str(row, c["etapa_neg"]),
            "id": get_str(row, c["id_conta"]),
            "d": get_iso(row, c["data"]),
        }
        if get_str(row, c["negocio"])
        else None
    ),
}

SCHEMAS["leadsConectado"] = {
    "filename_glob": "Leads por Etapa - Lead Conectado*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Etapa do Lead"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe"],
    "resolve_cols": lambda h: {
        "negocio": find_col(h, r"Nome do negócio"),
        "etapa_lead": find_col(h, r"Etapa do Lead"),
        "etapa_neg": find_col(h, r"Etapa do negócio"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "n": get_str(row, c["negocio"]),
            "el": get_str(row, c["etapa_lead"]),
        }
        if get_str(row, c["negocio"])
        else None
    ),
}

SCHEMAS["leadsAbordagem"] = {
    "filename_glob": "Leads por Etapa - Abordagem Iniciada*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Etapa do Lead"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe"],
    "resolve_cols": lambda h: {
        "negocio": find_col(h, r"Nome do negócio"),
        "etapa_lead": find_col(h, r"Etapa do Lead"),
        "etapa_neg": find_col(h, r"Etapa do negócio"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "n": get_str(row, c["negocio"]),
            "el": get_str(row, c["etapa_lead"]),
        }
        if get_str(row, c["negocio"])
        else None
    ),
}

SCHEMAS["oportunidades"] = {
    "filename_glob": "Oportunidades por empresa TATM*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Apres. Solução"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe", r"Fonte da qualificação", r"Carteira Macro"],
    "resolve_cols": lambda h: {
        "negocio": find_col(h, r"Nome do negócio"),
        "conta": find_col(h, r"Negócio: Conta"),
        "etapa": find_col(h, r"Etapa do negócio"),
        "data_apres": find_col(h, r"Data de entrada - Apres\. Solução"),
        "data_criacao": find_col(h, r"Data de criação"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "n": get_str(row, c["negocio"]),
            "c": get_str(row, c["conta"]),
            "e": get_str(row, c["etapa"]),
            "d": get_iso(row, c["data_criacao"]),
            "dap": get_iso(row, c["data_apres"]),
        }
        if get_str(row, c["negocio"])
        else None
    ),
}

SCHEMAS["negocios"] = {
    "filename_glob": "Negócios por xDR Plataforma*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Tamanho da Empresa"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe", r"Fonte da qualificação"],
    "resolve_cols": lambda h: {
        "negocio": find_col(h, r"Nome do negócio"),
        "conta": find_col(h, r"Negócio: Conta"),
        "etapa": find_col(h, r"Etapa do negócio"),
        "data": find_col(h, r"Data de criação"),
        "tamanho": find_col(h, r"Tamanho da Empresa"),
        "tipo": find_col(h, r"Tipo de negócio"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "n": get_str(row, c["negocio"]),
            "c": get_str(row, c["conta"]),
            "e": get_str(row, c["etapa"]),
            "d": get_iso(row, c["data"]),
            "tm": get_str(row, c["tamanho"]),
        }
        if get_str(row, c["negocio"])
        else None
    ),
}

SCHEMAS["bdrAe"] = {
    "filename_glob": "BDR AE*.xlsx",
    "header_tokens": ["Nome do membro da equipe", "Nome da conta"],
    "bdr_col_re": r"Nome do membro da equipe",
    "ffill_res": [r"Nome do membro da equipe", r"AE Principal", r"Propensão Cortex"],
    "resolve_cols": lambda h: {
        "conta": find_col(h, r"Nome da conta"),
        "carteira_macro": find_col(h, r"Carteira Macro"),
        "carteira_nova": find_col(h, r"Carteira Nova"),
        "papel": find_col(h, r"Papel na equipe"),
    },
    "extract": lambda row, c, cls: (
        {
            "b": cls["display"], "g": cls["grupo"],
            "c": get_str(row, c["conta"]),
            "cm": get_str(row, c["carteira_macro"]),
            "cn": get_str(row, c["carteira_nova"]),
        }
        if get_str(row, c["conta"])
        else None
    ),
}

# ============================================================
# CSV: MVP Fabbrica (lista TESTE)
# ============================================================
def parse_mvp_csv(path):
    with path.open(encoding="utf-8") as f:
        lines = f.readlines()
    header_idx = -1
    for i, line in enumerate(lines):
        if "Nome da conta" in line and "BDR da conta" in line:
            header_idx = i
            break
    if header_idx < 0:
        raise RuntimeError("MVP CSV: header não encontrado")
    reader = csv_mod.DictReader(lines[header_idx:])
    records = []
    for row in reader:
        nome = (row.get("Nome da conta") or "").strip()
        if not nome:
            continue
        bdr_raw = (row.get("BDR da conta") or "").strip()
        cls = classify(bdr_raw)
        records.append({
            "c": nome,
            "id": (row.get("ID da conta") or "").strip(),
            "br": bdr_raw,
            "bg": cls["grupo"] if cls else None,
            "bd": cls["display"] if cls else None,
        })
    return records

# ============================================================
# CSV: Alocadas Controle (lista CONTROLE)
# ============================================================
def parse_controle_csv(path):
    with path.open(encoding="utf-8") as f:
        reader = csv_mod.DictReader(f)
        records = []
        for row in reader:
            nome = (row.get("Nome da empresa") or "").strip()
            if not nome:
                continue
            records.append({
                "c": nome,
                "id": (row.get("AccountId") or "").strip(),
                "stage": (row.get("StageName") or "").strip(),
                "leadSource": (row.get("LeadSource") or "").strip(),
            })
        return records

# ============================================================
# RUN
# ============================================================
def find_file(glob):
    # Substring-fragment matching (avoids regex special-char issues like [Pre-Vendas])
    # Picks the MOST RECENTLY MODIFIED match — robust to duplicates like "file (1).csv"
    fragments = [f for f in glob.split("*") if f]
    matches = []
    for p in DATA_DIR.iterdir():
        if not p.is_file():
            continue
        name = p.name
        idx, ok = 0, True
        for frag in fragments:
            pos = name.find(frag, idx)
            if pos < 0:
                ok = False; break
            idx = pos + len(frag)
        if ok:
            matches.append(p)
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)

def main():
    out = {}
    for key, schema in SCHEMAS.items():
        path = find_file(schema["filename_glob"])
        if not path:
            print(f"  ✗ {key}: arquivo não encontrado ({schema['filename_glob']})", file=sys.stderr)
            continue
        try:
            recs = parse_xlsx(path, schema, debug=False)
            out[key] = {"fileName": path.name, "records": recs}
            print(f"  ✓ {key:<24} {len(recs):>6} registros · {path.name}")
        except Exception as e:
            print(f"  ✗ {key:<24} erro · {e}", file=sys.stderr)
            import traceback; traceback.print_exc()

    mvp_path = find_file("MVP Fabbrica*.csv")
    if mvp_path:
        try:
            recs = parse_mvp_csv(mvp_path)
            out["mvpFabbrica"] = {"fileName": mvp_path.name, "records": recs}
            print(f"  ✓ {'mvpFabbrica':<24} {len(recs):>6} registros · {mvp_path.name}")
        except Exception as e:
            print(f"  ✗ mvpFabbrica: erro · {e}", file=sys.stderr)

    ctrl_path = find_file("FGC_CS_GY_Alocadas*Criar_Negocios*.csv")
    if ctrl_path:
        try:
            recs = parse_controle_csv(ctrl_path)
            out["alocadasControle"] = {"fileName": ctrl_path.name, "records": recs}
            print(f"  ✓ {'alocadasControle':<24} {len(recs):>6} registros · {ctrl_path.name}")
        except Exception as e:
            print(f"  ✗ alocadasControle: erro · {e}", file=sys.stderr)

    # Extrai DATA do Dashboard ICP mais recente (operação do dia)
    saidas_dir = DATA_DIR / "saidas"
    icp_candidates = list(saidas_dir.glob("Dashboard_ICP_*.html")) if saidas_dir.exists() else []
    if icp_candidates:
        icp_html_path = max(icp_candidates, key=lambda p: p.stat().st_mtime)
        try:
            html = icp_html_path.read_text(encoding="utf-8")
            m = re.search(r'const DATA = (\{.*?\});\s*\n', html, re.DOTALL)
            if m:
                icp_data = json.loads(m.group(1))
                out["icp"] = icp_data
                print(f"  ✓ {'icp (operação dia)':<24} deals={len(icp_data.get('deals',[]))} carteira={len(icp_data.get('carteira',[]))} atv={len(icp_data.get('activities',[]))} opps={len(icp_data.get('opps',[]))} · {icp_html_path.name}")
            else:
                print(f"  ✗ icp: bloco DATA não encontrado em {icp_html_path.name}", file=sys.stderr)
        except Exception as e:
            print(f"  ✗ icp: erro · {e}", file=sys.stderr)
    else:
        print(f"  ✗ icp: nenhum Dashboard_ICP_*.html encontrado em {saidas_dir}", file=sys.stderr)

    OUT_PATH.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")))
    size_mb = OUT_PATH.stat().st_size / (1024 * 1024)
    print(f"\nGerado: {OUT_PATH} ({size_mb:.2f} MB)")

    # Inject JSON into piloto_fabbrica.html
    html_path = Path(__file__).parent / "piloto_fabbrica.html"
    if html_path.exists():
        html = html_path.read_text(encoding="utf-8")
        json_str = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
        new_html = re.sub(
            r'(<script id="preloaded-data" type="application/json">).*?(</script>)',
            lambda m: m.group(1) + json_str + m.group(2),
            html,
            count=1,
            flags=re.DOTALL,
        )
        html_path.write_text(new_html, encoding="utf-8")
        new_size_mb = html_path.stat().st_size / (1024 * 1024)
        print(f"Injetado em: {html_path} ({new_size_mb:.2f} MB)")

if __name__ == "__main__":
    main()
